import csv
from pathlib import Path

from openpyxl import load_workbook

MAX_ROWS = 200


def extract_spreadsheet(path: Path) -> str:
    if path.suffix.lower() == ".csv":
        return _csv_to_markdown(path)
    return _xlsx_to_markdown(path)


def _csv_to_markdown(path: Path) -> str:
    with path.open(newline="", encoding="utf-8", errors="replace") as f:
        rows = [[str(c) for c in row] for row in csv.reader(f)]
    return f"# {path.name}\n\n{_rows_to_table(rows)}"


def _xlsx_to_markdown(path: Path) -> str:
    wb = load_workbook(str(path), data_only=True, read_only=True)
    parts = [f"# {path.name}\n"]
    for ws in wb.worksheets:
        rows: list[list[str]] = []
        for row in ws.iter_rows(values_only=True):
            rows.append(["" if c is None else str(c) for c in row])
            if len(rows) > MAX_ROWS:
                break
        parts.append(f"\n## Sheet: {ws.title}\n\n{_rows_to_table(rows)}")
    return "\n".join(parts)


def _rows_to_table(rows: list[list[str]]) -> str:
    if not rows:
        return "_(empty)_\n"
    truncated = len(rows) > MAX_ROWS
    rows = rows[:MAX_ROWS]
    header, *body = rows
    ncols = len(header) or 1
    lines = ["| " + " | ".join(header) + " |", "|" + "---|" * ncols]
    for r in body:
        r = (r + [""] * ncols)[:ncols]
        lines.append("| " + " | ".join(r) + " |")
    if truncated:
        lines.append(f"\n_(truncated to first {MAX_ROWS} rows)_")
    return "\n".join(lines) + "\n"
